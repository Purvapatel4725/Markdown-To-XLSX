import openpyxl
import re

def parse_markdown_table(md_table):
    """
    Parse a markdown table and return the headers and rows.
    """
    # Split the table into lines
    lines = md_table.strip().split("\n")

    # Extract headers and rows
    headers = [col.strip() for col in lines[0].split("|") if col.strip()]
    rows = [
        [col.strip() for col in line.split("|") if col.strip()]
        for line in lines[2:]  # Skipping the separator line
    ]

    return headers, rows

def create_excel_from_markdown(md_table, output_file):
    """
    Create an Excel file from a markdown table.
    """
    # Parse the markdown table
    headers, rows = parse_markdown_table(md_table)

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add the headers to the first row
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # Add the rows
    for row_num, row in enumerate(rows, start=2):
        for col_num, cell_value in enumerate(row, start=1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    # Save the workbook
    workbook.save(output_file)
    print(f"Excel file created: {output_file}")

# Sample markdown table
markdown_table = """
| Protocol | Network                      | Metric | Next Hop                   | Interface |
|----------|------------------------------|--------|----------------------------|-----------|
| S        | ::/0                         | 1/0    | 2001:DB8:CAFE:99::D1       | -         |
| C        | 2001:DB8:CAFE:99::/64        | 0/0    | -                          | Vlan99    |
| L        | 2001:DB8:CAFE:99::A1/128     | 0/0    | -                          | Vlan99    |
| C        | 2001:DB8:CAFE:110::/64       | 0/0    | -                          | Vlan110   |
| L        | 2001:DB8:CAFE:110::A1/128    | 0/0    | -                          | Vlan110   |
| C        | 2001:DB8:CAFE:120::/64       | 0/0    | -                          | Vlan120   |
| L        | 2001:DB8:CAFE:120::A1/128    | 0/0    | -                          | Vlan120   |
| C        | 2001:DB8:CAFE:200::/64       | 0/0    | -                          | Vlan200   |
| L        | 2001:DB8:CAFE:200::A1/128    | 0/0    | -                          | Vlan200   |
| L        | FF00::/8                     | 0/0    | -                          | Null0     |


"""

# Create the Excel file from the markdown table
create_excel_from_markdown(markdown_table, "output_table.xlsx")
